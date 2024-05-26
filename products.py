import openpyxl as op
import json
from time import sleep
from tqdm import tqdm
from datetime import datetime, timedelta
from my_modules.mp_apiconnect.moy_sklad import ms
from my_modules.mp_apiconnect.ozon import ozon
from my_modules.log.log import log

class Product:
    product_list = []
    statistic_data = {} # кэш для выгрузки статистики по дням

    def __init__(self, title, changed, skus, skus_oz, price, prev_change, prev_profit_day,
                 selfcost, marj, orders_total, items_total):
        self.title = title
        self.changed = changed
        self.skus = skus
        self.skus_oz = skus_oz
        self.price = price
        self.prev_change = prev_change
        self.prev_profit_day = prev_profit_day
        self.prev_selfcost = selfcost
        self.selfcost = selfcost
        timepassed = datetime.now() - changed
        self.days_passed = timepassed.days + timepassed.seconds / (24 * 3600)
        self.new_change = None
        self.profit_day = None
        self.marj = marj
        self.profit = None
        self.new_price = None
        self.pcs_ordered = None
        self.orders = 0
        self.orders_total = orders_total
        self.items_total = items_total
        self.on_stock = 0

    def __str__(self):
        return self.title

    def display(self):
        print(self.__dict__)

    @classmethod
    def add_stock(cls):
        """
        Добавляет в данные каждой группы продуктов поле '.on_stock' - количество товара на складе
        """
        stock_loaded = ozon.get_ozon_stock()

        stock_data = {}
        for item in stock_loaded:
            sku = item['item_code']
            stock_data[sku] = stock_data.get(sku, 0) + item['free_to_sell_amount']

        for group in cls.product_list:
            for sku in group.skus:
                if sku in stock_data:
                    group.on_stock += stock_data[sku]

    @classmethod
    def add_actions(cls) -> None:
        """
        Добавляет атрибут 'in_action'. True, если хотя бы один товар группы участвует в акции
        """
        for product in cls.product_list:
            in_action = False
            for sku in product.skus:
                product_data = ozon.product_info(sku)
                if product_data['in_action']:
                    in_action = True
                    break
            product.in_action = in_action

    def if_badsellers_himarj(self, period: int, sale_limit: int, marj_limit: float):
        """
        Определяем товар, который плохо продается, но высокая маржа
        :param sale_limit: штук в мес. Если продано меньше, цена снижается
        :param marj_limit: Минимальная маржа, при которой правило действует, напр.:
            0 - продажа по себестоимости
            0.3 - продажа с маржой 30%
        :return: True, если надо снизить цену
        """

        return (
            # Дней без изменений > period
            self.days_passed > period and
            # Маржа не указана либо выше лимита
            (self.marj is None or self.marj > marj_limit) and
            # Продано меньше лимита
            self.pcs_ordered < sale_limit and
            # Товар есть на складе
            self.on_stock
        )


    def count_total(self):
        """
        Добавляет количество заказов и штук за период к общему количеству. Общее количество используется в модуле
        формирования заказа для подсчета среднего кол-ва товара в заказе.
        :return:
        """
        self.orders_total += self.orders
        self.items_total += self.pcs_ordered

    @classmethod
    def get_ozon_dinamic_price_data(cls):
        """ Создаем лист товаров на основе записи в last_change.json"""
        with open ('data/last_change.json', encoding='utf-8') as last_change_json:
            last_change_data = json.load(last_change_json)

        for prod in last_change_data:  # создаем новый Product и добавляем его в список
            new_product_gr = Product(
                title = prod['title'],
                changed = datetime.fromisoformat(prod['changed']), # время последнего изменения
                skus = prod['skus'],                               # список наших sku
                skus_oz = prod['skus_oz'],                         # список sku озона
                price = prod['price'],                             # текущая цена
                prev_change = prod['prev_change'],                 # предыдущее изменение цены (1 или -1)
                prev_profit_day = prod['prev_profit_day'],         # прибыль в день за прошлый период
                selfcost = prod['selfcost'],                       # себестоимость средняя по всем товарам
                marj = prod.get('marj', None),                     # маржа
                orders_total = prod.get('orders_total', 0),        # заказов за все время (для модуля формирования заказа)
                items_total=prod.get('items_total', 0),            # штук за все время (для модуля формирования заказа)
            )

            cls.product_list.append(new_product_gr)

        return cls.product_list


    @classmethod
    def append_new_products(cls):
        """
        Функция дополняет список товаров товарами из файла new_products.xls и очищает файл
        """

        workbook = op.load_workbook('data/new_products.xlsx')
        sheet = workbook.active
        date_from = datetime.now() - timedelta(days=14)

        added_skus = []                     # создаем список уже занесенных артикулов для последующей проверки новых
        for product in Product.product_list:
            added_skus.extend(product.skus)

        # создаем новый товар с полями из Excel
        row = 2
        while sheet.cell(row, 1).value:     # перебор строк, пока значение в поле артикул заполнено
            new_product = cls(
                title=sheet.cell(row, 2).value,
                skus=sheet.cell(row, 1).value.strip().split(', '),
                skus_oz=[],
                price=int(sheet.cell(row, 3).value),
                prev_change=int(sheet.cell(row, 4).value),
                selfcost=None,
                changed=date_from,
                prev_profit_day=0,
                marj=0,
                orders_total=0,
                items_total=0
                )

            # проверяем, что новых ску нет в базе
            pass_sku = False
            for sku in new_product.skus:
                if sku in added_skus:
                    log.add(f'[!] Артикул {sku} {new_product} уже есть в базе. Товар пропущен.')
                    pass_sku = True
            if pass_sku:
                row += 1
                continue

            # добавляем sku Озона
            skus_oz = ozon.get_ozon_skus(new_product.skus)
            if skus_oz is None:
                log.add(f'[!] Ошибка при поиске артикулов нового товара {new_product}: {new_product.skus}'
                        f'Товар пропущен.')
                row += 1
                continue
            new_product.skus_oz = skus_oz

            # добавляем новый продукт в базу
            cls.product_list.append(new_product)
            log.add(f'[a] Товар {new_product} добавлен')
            row += 1

        while True:
            try:
                workbook.save('data/new_products.xlsx')
                break
            except Exception as e:
                log.add(f'[!] Ошибка сохранения файла new_products.xlsx. Закройте файл, если он открыт. {e}')
                sleep(3)


    @classmethod
    def clear_new_products_file(cls):
        """
        Функция удаляет строки с новыми продуктами из файла new_products.xlsx
        """

        workbook = op.load_workbook('data/new_products.xlsx')
        sheet = workbook.active
        row = 2
        while sheet.cell(row, 1).value:  # перебор строк, пока значение в поле артикул заполнено
            sheet.delete_rows(row)
        workbook.save('data/new_products.xlsx')


    @classmethod
    def save_ozon_dinamic_price_data(cls):
        """
        Сохраняем измененные данные в файл .json
        """
        new_json = []
        for product in cls.product_list:
            if product.new_price is None:
                data = {
                    "title": product.title,
                    "changed": product.changed.isoformat(),
                    "skus": product.skus,
                    "skus_oz": product.skus_oz,
                    "price": product.price,
                    "prev_change": product.prev_change,
                    "prev_profit_day": product.prev_profit_day,
                    "selfcost": product.selfcost,
                    "marj": product.marj,
                    "orders_total": product.orders_total,
                    "items_total": product.items_total
                }
            else:
                data = {
                    "title": product.title,
                    "changed": datetime.now().isoformat(),
                    "skus": product.skus,
                    "skus_oz": product.skus_oz,
                    "price": product.new_price,
                    "prev_change": product.new_change,
                    "prev_profit_day": product.profit_day,
                    "selfcost": product.selfcost,
                    "marj": product.marj,
                    "orders_total": product.orders_total,
                    "items_total": product.items_total
                }
            new_json.append(data)

        while True:
            try:
                with open ('data/last_change.json', mode='w', encoding='utf-8') as last_change_json:
                    json.dump(new_json, last_change_json, ensure_ascii=False)
                break
            except Exception as e:
                log.add(f'[!] Ошибка сохранения json файла. Закройте файл, если он открыт. {e}')
                sleep(3)

    @classmethod
    def add_new_selfcost(cls):
        """
        Функция добавляет новую себестоимость из Моего Склада
        """
        ms_selfcost = ms.get_selfcosts()  # получаем словарь с себестоиомстью из МС

        for product in cls.product_list:
            count_skus = 0
            count_selfcost = 0

            for sku in product.skus:
                if sku in ms_selfcost and ms_selfcost[sku]:
                    count_skus += 1
                    count_selfcost += ms_selfcost[sku]

            if count_selfcost:
                product.selfcost = count_selfcost / count_skus
            else:
                print(f"""
                У товара {product.title} (SKU {product.skus}) нет данных о себестоимости.
                Предыдущая себестоимость: {product.prev_selfcost}
                """)
                product.selfcost = product.prev_selfcost

            if product.prev_selfcost:
                selfcost_change = 1 - product.prev_selfcost / product.selfcost
                if abs (selfcost_change) > 0.2:
                    print(f'[i]Себестоимость товара {sku} {ozon.get_name(sku)} изменилась на {round(selfcost_change * 100)}%')


    @classmethod
    def save_changes_xls(cls):
        """
        Сохраняем изменения в журнал xls
        """

        workbook = op.load_workbook('data/changes.xlsx')
        sheet = workbook.active

        for product in cls.product_list:
            if product.new_price is None: continue

            profit = product.profit if product.profit else 0
            new_row = [
                datetime.today().strftime("%d.%m.%Y"),
                round(product.days_passed),
                ' '.join(product.skus),
                product.title,
                product.price,
                product.new_price,
                product.new_change,
                ms.products[product.skus[0]]['price'],
                product.pcs_ordered,
                product.orders,
                round(product.pcs_ordered / product.days_passed, 1),
                round(product.profit_day),
                round(profit),
                product.marj,
                round(product.selfcost),
            ]
            if product.prev_selfcost:
                new_row.append(round(product.prev_selfcost))
                new_row.append(round(100 * (1 - product.selfcost / product.prev_selfcost)))

            sheet.append(new_row)

        while True:
            try:
                workbook.save('data/changes.xlsx')
                break
            except Exception as e:
                log.add(f'[!] Ошибка сохранения xls файла. Закройте файл, если он открыт. {e}')
                sleep(3)

    def is_discount(self) -> bool:
        """
        Проверяет, участвует ли хоть один из товаров группы в акциях
        (цена товара на площадке не соотвествует последней сохраненной цене)
        :return:  True если участвует
        """
        saved_price = self.price

        for sku in self.skus:
            real_price_str = ozon.product_info(sku)['price']
            real_price = int(real_price_str.split('.')[0])
            print(real_price, saved_price)
            if real_price != saved_price:
                return True

        return False

    def add_oz_ordered(self):
        """Дополняем товар данными о количестве заказанных товаров на озон за период"""

        date_from_str = self.changed.strftime('%Y-%m-%d')

        # Если статистика за этот период уже загружалась, берем из кэша
        if date_from_str in Product.statistic_data:
            all_data = Product.statistic_data[date_from_str]
            print('Данные о заказах взяты из кэша')

        # Если в кэше нет, загружаем и сохраняем в кэш
        else:
            all_data = ozon.get_ozon_orders_statistic(self.changed, datetime.now())
            Product.statistic_data[date_from_str] = all_data

        pcs_ordered = 0
        for product_oz_sku in self.skus_oz:
            for prod_stat in all_data:
                if product_oz_sku == prod_stat['sku_oz']:
                    pcs_ordered += prod_stat['pcs_ordered']
        self.pcs_ordered = pcs_ordered

    def add_orders_amount(self):
        """
        Дополняем товар количеством заказов за период (необходимо для более корректного подсчета прибыльности,
        т.к. считаем не кол-во товаров, а кол-во заказов, чтобы оптовые заказы не влияли на результат)
        """
        orders = 0
        for sku_oz in self.skus_oz:
            orders += len(ozon.get_orders_by_product(sku_oz, self.changed, status='not_canceled')['orders'])

        self.orders = orders

    def add_profit(self, minimum_transactions, aquiring, profit_koef):
        """Дополняем товар средним доходом"""

        #1. Собираем номера отправлений доставленных заказов
        #TODO: добавить цикл для подсчета всех товаров в списке, а не только первого
        postings = []
        for sku_oz in self.skus_oz:
            postings.extend(ozon.get_orders_by_product(sku_oz, self.changed, status = 'delivered')['postings'])

        #2. Считаем средний доход от доставленных товаров
        postings_total = 0  # кол-во отправлений
        income_total = 0    # сумма перечисленная озоном
        for posting in tqdm(postings):
            income = ozon.get_posting_income(posting)
            if income is None: continue # если в отправлении были разные товары, его не учитываем

            income_total += income
            postings_total += 1

        if postings_total < minimum_transactions: # пропускаем товары, у которых мало доставленных заказов
            self.profit = None
        else:
            profit = income_total / postings_total - self.selfcost - aquiring * self.price
            profit = profit * profit_koef
            self.profit = round(profit, ndigits=2)

            self.marj = round(profit / self.selfcost, 2)  # маржинальность

    def count_profit_per_day(self):
        """
        Считаем среднюю прибыль в день. Экстраполируем данные из транзакций на все товары и делим на прошедшие дни
        За расчет берем количество заказов, а не товаров, чтобы оптовые заказы не влияли на результат
        """
        profit = self.profit * self.orders
        self.profit_day = profit / self.days_passed

    def count_new_price(self, koef, marj_min, profit_min, forced_down: bool):

        # Если принудительное снижение
        if forced_down:
            self.new_change = -1

        # Если нет принудительного снижения
        else:
            if self.profit_day > self.prev_profit_day:
                self.new_change = self.prev_change
            else:
                self.new_change = -self.prev_change

            # Если цена была ниже цены Zeero, повышаем цену
            if self.price < ms.products[self.skus[0]]['price']:
                self.new_change = 1
                log.add(f'[i] Цена товара {self.title} ниже допустимой')

            # Если маржинальность опустилась ниже допустимой, повышаем цену
            if self.marj < marj_min:
                self.new_change = 1
                log.add(f'[i] У товара {self.title} маржинальность {self.marj} ниже допустимой {marj_min}')

            # Если прибыль с единицы ниже допустимой, повышаем цену
            if self.profit < profit_min:
                self.new_change = 1
                log.add(f'[i] У товара {self.title} прибыльность {self.profit} ниже допустимой {profit_min}')

        self.new_price = round(self.price + self.price * koef * self.new_change)

        if not self.prev_profit_day: self.prev_profit_day = 0
        if not self.profit_day: self.profit_day = 0
        log.add(f'[i] Прибыль в день. Была: {round(self.prev_profit_day)}, стала: {round(self.profit_day)}')
        log.add(f'[i] Цена повышена' if self.new_change == 1 else f'[i] Цена понижена')
        log.add(f'[i] Старая цена: {self.price}. Новая цена: {self.new_price}.')

    def change_price_oz(self):
        data_list = []
        for sku in self.skus:
            data = {'sku': sku, 'price':self.new_price}
            data_list.append(data)
        ozon.update_prices_ozon(data_list)


